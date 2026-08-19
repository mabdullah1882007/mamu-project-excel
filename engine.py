import openpyxl
import json
import re
import os
import sqlite3
from datetime import datetime
from collections import defaultdict
from io import BytesIO

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
MAPPINGS_FILE = os.path.join(BASE_DIR, "mappings.json")
HISTORY_DB = os.path.join(BASE_DIR, "data", "history.db")

os.makedirs(os.path.join(BASE_DIR, "data"), exist_ok=True)


def get_db():
    conn = sqlite3.connect(HISTORY_DB)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS history (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            filename TEXT,
            timestamp TEXT,
            total_sheet2_rows INTEGER,
            matched_rows INTEGER,
            unmatched_rows INTEGER,
            total_purchase_qty REAL,
            total_purchase_value REAL,
            total_purchase_tax REAL
        )
    """)
    conn.commit()
    return conn


def save_history(filename, total, matched, unmatched, qty, value, tax):
    conn = get_db()
    conn.execute(
        "INSERT INTO history VALUES (NULL,?,?,?,?,?,?,?,?)",
        (filename, datetime.now().isoformat(), total, matched, unmatched, qty, value, tax),
    )
    conn.commit()
    conn.close()


def load_history():
    conn = get_db()
    rows = conn.execute("SELECT * FROM history ORDER BY id DESC").fetchall()
    conn.close()
    return rows


def load_mappings():
    if os.path.exists(MAPPINGS_FILE):
        with open(MAPPINGS_FILE, "r") as f:
            return json.load(f)
    return {}


def save_mappings(mappings):
    with open(MAPPINGS_FILE, "w") as f:
        json.dump(mappings, f, indent=2)


def normalize_hs(hs_string):
    """Normalize HS code by stripping trailing zeros: '5802.10' -> '5802.1', '2716.00' -> '2716'"""
    if hs_string is None:
        return ""
    s = str(hs_string).strip()
    try:
        val = float(s)
        if val == int(val):
            return str(int(val))
        formatted = f"{val:g}"
        return formatted
    except ValueError:
        return s


def extract_numeric_hs(hs_string):
    if hs_string is None:
        return None
    hs_str = str(hs_string).strip()
    match = re.match(r"(\d{4}\.\d+)", hs_str)
    if match:
        return normalize_hs(match.group(1))
    match = re.match(r"(\d+)", hs_str)
    if match:
        return match.group(1)
    return normalize_hs(hs_str)


def parse_formula_value(val):
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    if s.startswith("="):
        expr = s[1:].strip()
        try:
            return float(expr)
        except ValueError:
            return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0


def parse_sheet2(wb, mappings):
    ws = wb["Sheet2"]

    headers = {}
    for c in range(1, ws.max_column + 1):
        val = ws.cell(1, c).value
        if val:
            headers[str(val).strip().upper()] = c

    hs_col = headers.get("HS CODE", 10)
    type_col = headers.get("TYPE", 11)
    qty_col = headers.get("QUANTITY", 14)
    uom_col = headers.get("UOM", 15)
    value_col = headers.get("VALUE OF PURCHASES EXCLUDING SALES TAX", 16)
    st_col = headers.get("SALES TAX/ FED IN ST MODE", 17)
    rate_col = headers.get("RATE", 13)

    aggregated = defaultdict(lambda: {"qty": 0.0, "value": 0.0, "st": 0.0, "uom": "", "rows": []})

    for r in range(2, ws.max_row + 1):
        hs_raw = ws.cell(r, hs_col).value
        type_raw = ws.cell(r, type_col).value
        qty_raw = ws.cell(r, qty_col).value
        value_raw = ws.cell(r, value_col).value
        st_raw = ws.cell(r, st_col).value
        uom_raw = ws.cell(r, uom_col).value
        rate_raw = ws.cell(r, rate_col).value

        if not hs_raw and not type_raw:
            continue

        qty_val = parse_formula_value(qty_raw)
        value_val = parse_formula_value(value_raw)

        # For tax: try to evaluate =SUM(P*M%) formula as value * rate / 100
        st_val = parse_formula_value(st_raw)
        if st_val == 0 and value_val > 0:
            st_str = str(st_raw).strip() if st_raw else ""
            if "SUM" in st_str.upper() and "%" in st_str:
                rate = parse_formula_value(rate_raw)
                if rate > 0:
                    st_val = value_val * rate / 100

        if qty_val == 0 and value_val == 0 and st_val == 0:
            continue

        type_str = str(type_raw).strip().upper() if type_raw else ""
        description = mappings.get(type_str, type_str)
        hs_numeric = extract_numeric_hs(hs_raw) if hs_raw else ""

        key = (hs_numeric, description)
        aggregated[key]["qty"] += qty_val
        aggregated[key]["value"] += value_val
        aggregated[key]["st"] += st_val
        if uom_raw:
            aggregated[key]["uom"] = str(uom_raw).strip()
        aggregated[key]["rows"].append(r)

    return aggregated


def find_h_data_rows(ws):
    data_rows = []
    for r in range(6, ws.max_row + 1):
        desc = ws.cell(r, 1).value
        if desc and str(desc).strip() not in ("Total", "total", ""):
            data_rows.append(r)
    return data_rows


def process_excel(uploaded_bytes, mappings):
    wb = openpyxl.load_workbook(BytesIO(uploaded_bytes), data_only=False)

    if "Sheet2" not in wb.sheetnames:
        return None, {"error": "Sheet 'Sheet2' not found in workbook"}
    if "H" not in wb.sheetnames:
        return None, {"error": "Sheet 'H' not found in workbook"}

    aggregated = parse_sheet2(wb, mappings)
    ws_h = wb["H"]
    h_data_rows = find_h_data_rows(ws_h)

    h_lookup = {}
    for r in h_data_rows:
        desc = str(ws_h.cell(r, 1).value).strip()
        hs_raw = ws_h.cell(r, 2).value
        if hs_raw is not None:
            hs_num = normalize_hs(str(hs_raw))
        else:
            hs_num = ""
        h_lookup[(hs_num, desc)] = r

    matched = 0
    unmatched_details = []
    matched_details = []
    total_qty = 0.0
    total_value = 0.0
    total_st = 0.0

    for (hs_num, desc), data in aggregated.items():
        h_row = h_lookup.get((hs_num, desc))

        if h_row is not None:
            ws_h.cell(h_row, 7).value = round(data["qty"], 4)
            ws_h.cell(h_row, 8).value = round(data["value"], 2)
            ws_h.cell(h_row, 9).value = round(data["st"], 2)
            matched += 1
            total_qty += data["qty"]
            total_value += data["value"]
            total_st += data["st"]
            matched_details.append({
                "h_row": h_row,
                "description": desc,
                "hs_code": hs_num,
                "qty": data["qty"],
                "value": data["value"],
                "tax": data["st"],
                "source_rows": data["rows"],
            })
        else:
            unmatched_details.append({
                "hs_code": hs_num,
                "description": desc,
                "qty": data["qty"],
                "value": data["value"],
                "tax": data["st"],
                "source_rows": data["rows"],
            })

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    report = {
        "total_sheet2_groups": len(aggregated),
        "h_data_rows": len(h_data_rows),
        "matched": matched,
        "unmatched": len(unmatched_details),
        "total_qty": total_qty,
        "total_value": total_value,
        "total_st": total_st,
        "matched_details": matched_details,
        "unmatched_details": unmatched_details,
    }

    return output, report
