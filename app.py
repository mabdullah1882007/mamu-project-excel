import streamlit as st
import pandas as pd
import openpyxl
import os
from io import BytesIO

from engine import (
    process_excel, load_mappings, save_mappings,
    load_history, save_history, find_h_data_rows, extract_numeric_hs,
    parse_formula_value, parse_sheet2
)

st.set_page_config(page_title="Purchase Data Linker", page_icon="📊", layout="wide")

# ── Sidebar ───────────────────────────────────────────────────────────────────

st.sidebar.title("📊 Purchase Linker")
page = st.sidebar.radio("Navigate", ["Upload & Process", "Configuration", "History", "Report"])

st.sidebar.markdown("---")
st.sidebar.markdown("#### How it works")
st.sidebar.markdown("""
1. Upload Excel with **H** and **Sheet2** tabs
2. System links purchases by **HS Code + TYPE**
3. Fills Purchase columns in H tab
4. Download the processed file
""")


# ── Helper: parse formula for preview display ─────────────────────────────────

def parse_preview_value(val):
    """Convert formula strings to readable values for preview."""
    if val is None:
        return ""
    if isinstance(val, (int, float)):
        return val
    s = str(val).strip()
    if s.startswith("="):
        return parse_formula_value(s)
    try:
        return float(s)
    except ValueError:
        return s


def parse_sheet2_preview(ws):
    """Parse Sheet2 showing readable values."""
    rows = []
    for r in range(1, min(ws.max_row + 1, 12)):
        row = []
        for c in range(1, min(ws.max_column + 1, 22)):
            val = ws.cell(r, c).value
            if r == 1:
                row.append(str(val or ""))
            else:
                row.append(str(parse_preview_value(val)))
        rows.append(row)
    return rows


# ── Page: Upload & Process ────────────────────────────────────────────────────

if page == "Upload & Process":
    st.title("Upload & Process Excel")

    uploaded = st.file_uploader("Upload Excel file", type=["xlsx"])

    if uploaded:
        file_bytes = uploaded.getvalue()
        st.success(f"Loaded: **{uploaded.name}** ({len(file_bytes):,} bytes)")

        tab1, tab2, tab3 = st.tabs(["Preview Sheet2", "Preview H Sheet", "Parsed Sheet2 Data"])

        with tab1:
            try:
                wb_preview = openpyxl.load_workbook(BytesIO(file_bytes), data_only=False)
                ws2 = wb_preview["Sheet2"]
                preview_rows = parse_sheet2_preview(ws2)
                cols = [f"C{c}" for c in range(1, min(ws2.max_column + 1, 22))]
                st.dataframe(pd.DataFrame(preview_rows, columns=cols), width="stretch")
                st.caption(f"Total rows: {ws2.max_row} — Formulas converted to values for display")
            except Exception as e:
                st.error(f"Error reading Sheet2: {e}")

        with tab2:
            try:
                ws_h = wb_preview["H"]
                preview_data = []
                for r in range(1, min(10, ws_h.max_row + 1)):
                    row = [str(ws_h.cell(r, c).value or "") for c in range(1, min(10, ws_h.max_column + 1))]
                    preview_data.append(row)
                cols_h = [str(ws_h.cell(1, c).value or f"C{c}") for c in range(1, min(10, ws_h.max_column + 1))]
                st.dataframe(pd.DataFrame(preview_data, columns=cols_h), width="stretch")
            except Exception as e:
                st.error(f"Error reading H sheet: {e}")

        with tab3:
            try:
                mappings = load_mappings()
                wb_preview2 = openpyxl.load_workbook(BytesIO(file_bytes), data_only=False)
                aggregated = parse_sheet2(wb_preview2, mappings)

                st.markdown(f"**{len(aggregated)} unique HS Code + TYPE groups found in Sheet2**")

                agg_data = []
                for (hs, desc), data in sorted(aggregated.items()):
                    agg_data.append({
                        "HS Code": hs,
                        "Description (TYPE)": desc,
                        "Qty": round(data["qty"], 2),
                        "Value": round(data["value"], 0),
                        "Tax": round(data["st"], 0),
                        "Source Rows": len(data["rows"]),
                    })

                if agg_data:
                    agg_df = pd.DataFrame(agg_data)
                    st.dataframe(agg_df, width="stretch")

                    total_qty = sum(d["Qty"] for d in agg_data)
                    total_val = sum(d["Value"] for d in agg_data)
                    total_tax = sum(d["Tax"] for d in agg_data)

                    c1, c2, c3 = st.columns(3)
                    c1.metric("Total Qty", f"{total_qty:,.2f}")
                    c2.metric("Total Value", f"{total_val:,.0f}")
                    c3.metric("Total Tax", f"{total_tax:,.0f}")
            except Exception as e:
                st.error(f"Error parsing Sheet2: {e}")

        st.markdown("---")

        # ── Process button ──────────────────────────────────────────────
        if st.button("Process & Link Purchases", type="primary", width="stretch"):
            mappings = load_mappings()
            with st.spinner("Processing..."):
                output, report = process_excel(file_bytes, mappings)

            if output is None:
                st.error(report["error"])
            else:
                st.session_state["output"] = output
                st.session_state["report"] = report
                st.session_state["filename"] = uploaded.name
                st.session_state["processed"] = True

        # ── Display results from session_state (persists across reruns) ─
        if st.session_state.get("processed"):
            report = st.session_state["report"]
            output = st.session_state["output"]

            st.success("Processing complete!")

            c1, c2, c3, c4 = st.columns(4)
            c1.metric("Matched Groups", report["matched"])
            c2.metric("Unmatched Groups", report["unmatched"])
            c3.metric("Total Purchase Qty", f"{report['total_qty']:,.2f}")
            c4.metric("Total Purchase Value", f"{report['total_value']:,.0f}")

            c5, c6 = st.columns(2)
            c5.metric("Total Purchase Tax", f"{report['total_st']:,.0f}")
            c6.metric("H Sheet Data Rows", report["h_data_rows"])

            if report["matched_details"]:
                st.subheader("Matched Items")
                match_df = pd.DataFrame([
                    {
                        "H Row": m["h_row"],
                        "Description": m["description"],
                        "HS Code": m["hs_code"],
                        "Qty": round(m["qty"], 2),
                        "Value": round(m["value"], 0),
                        "Tax": round(m["tax"], 0),
                        "Source Rows": len(m["source_rows"]),
                    }
                    for m in report["matched_details"]
                ])
                st.dataframe(match_df, width="stretch")

            if report["unmatched_details"]:
                st.subheader("Unmatched Items (Warning)")
                unmatch_df = pd.DataFrame([
                    {
                        "HS Code": u["hs_code"],
                        "Description": u["description"],
                        "Qty": round(u["qty"], 2),
                        "Value": round(u["value"], 0),
                        "Tax": round(u["tax"], 0),
                        "Source Rows": len(u["source_rows"]),
                    }
                    for u in report["unmatched_details"]
                ])
                st.dataframe(unmatch_df, width="stretch")

            st.markdown("---")
            st.download_button(
                "Download Processed File",
                data=output.getvalue(),
                file_name=f"{st.session_state.get('filename', 'output').replace('.xlsx', '')}_linked.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                width="stretch",
            )

            save_history(
                st.session_state.get("filename", "unknown"),
                report["total_sheet2_groups"],
                report["matched"],
                report["unmatched"],
                report["total_qty"],
                report["total_value"],
                report["total_st"],
            )

    else:
        # Clear session state when no file is uploaded
        for key in ["processed", "output", "report", "filename"]:
            st.session_state.pop(key, None)

        st.info("Upload an Excel file with **H** and **Sheet2** tabs to begin.")


# ── Page: Configuration ───────────────────────────────────────────────────────

elif page == "Configuration":
    st.title("Mapping Configuration")
    st.markdown("Map Sheet2 **TYPE** values to H sheet **Description** values.")

    mappings = load_mappings()

    st.subheader("Current Mappings")

    edited = st.data_editor(
        pd.DataFrame(list(mappings.items()), columns=["TYPE (Sheet2)", "Description (H Sheet)"]),
        num_rows="dynamic",
        width="stretch",
        key="mapping_editor",
    )

    if st.button("Save Mappings", type="primary"):
        new_map = {row["TYPE (Sheet2)"].upper(): row["Description (H Sheet)"]
                   for _, row in edited.iterrows()
                   if row["TYPE (Sheet2)"] and row["Description (H Sheet)"]}
        save_mappings(new_map)
        st.success("Mappings saved!")
        st.rerun()


# ── Page: History ─────────────────────────────────────────────────────────────

elif page == "History":
    st.title("Processing History")

    history = load_history()

    if not history:
        st.info("No processing history yet. Process a file first.")
    else:
        df = pd.DataFrame(
            history,
            columns=["ID", "Filename", "Timestamp", "Total Groups", "Matched", "Unmatched", "Qty", "Value", "Tax"],
        )
        df["Timestamp"] = pd.to_datetime(df["Timestamp"]).dt.strftime("%Y-%m-%d %H:%M")
        st.dataframe(df.drop(columns=["ID"]), width="stretch")

        st.markdown("---")
        st.subheader("Summary Stats")
        c1, c2, c3 = st.columns(3)
        c1.metric("Total Runs", len(df))
        c2.metric("Avg Match Rate", f"{df['Matched'].mean() / df['Total Groups'].mean() * 100:.1f}%")
        c3.metric("Total Value Processed", f"{df['Value'].sum():,.0f}")


# ── Page: Report ──────────────────────────────────────────────────────────────

elif page == "Report":
    st.title("Detailed Report")

    uploaded_report = st.file_uploader("Upload processed file for report", type=["xlsx"], key="report_upload")

    if uploaded_report:
        file_bytes = uploaded_report.getvalue()
        wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
        ws_h = wb["H"]

        st.subheader("H Sheet - Purchase Data Summary")

        data_rows = find_h_data_rows(ws_h)
        report_data = []
        total_qty = 0
        total_value = 0
        total_tax = 0

        for r in data_rows:
            desc = ws_h.cell(r, 1).value or ""
            hs = ws_h.cell(r, 2).value or ""
            uom = ws_h.cell(r, 3).value or ""
            pur_qty = ws_h.cell(r, 7).value or 0
            pur_val = ws_h.cell(r, 8).value or 0
            pur_st = ws_h.cell(r, 9).value or 0

            total_qty += pur_qty
            total_value += pur_val
            total_tax += pur_st

            report_data.append({
                "Row": r,
                "Description": str(desc),
                "HS Code": str(hs),
                "UoM": str(uom),
                "Purchase Qty": float(pur_qty),
                "Purchase Value": float(pur_val),
                "Purchase Tax": float(pur_st),
                "Has Data": "Yes" if pur_qty else "No",
            })

        df_report = pd.DataFrame(report_data)
        st.dataframe(df_report, width="stretch")

        st.subheader("Totals")
        c1, c2, c3 = st.columns(3)
        c1.metric("Total Purchase Qty", f"{total_qty:,.2f}")
        c2.metric("Total Purchase Value", f"{total_value:,.0f}")
        c3.metric("Total Purchase Tax", f"{total_tax:,.0f}")

        rows_with = sum(1 for r in report_data if r["Has Data"] == "Yes")
        rows_without = sum(1 for r in report_data if r["Has Data"] == "No")

        st.subheader("Coverage")
        st.progress(rows_with / len(report_data) if report_data else 0)
        st.markdown(f"**{rows_with}** of **{len(report_data)}** rows have purchase data filled")
        if rows_without > 0:
            st.warning(f"**{rows_without}** rows have no purchase data")
